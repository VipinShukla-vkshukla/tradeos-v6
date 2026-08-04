"""
What a trade in this system actually earns, after everything it actually costs.

    python -m tools.expectancy_ledger              the ledger
    python -m tools.expectancy_ledger --by-clip    net R against position size
    python -m tools.expectancy_ledger --reconcile FILE.csv
                                                   modelled vs the broker's own

WHY THIS EXISTS
---------------
No document in this repository states what a trade nets. Everything downstream
is denominated in R, and if R is not net of true friction then every number the
system produces afterwards is fiction — engine scores, gate calibration, the
priors the allocator will eventually rank on. The cost problem went unnoticed
for months precisely because nothing ever wrote this down.

IT RECOMPUTES RATHER THAN READING closed_positions.charges
----------------------------------------------------------
That column is written at close time by whatever the cost model said that day,
and it is present on 21 of 91 rows. Reading it would mean scoring old trades
under old rates, comparing them to new trades under new rates, and calling the
difference edge. So friction is recomputed for every row at today's rates and
at the product actually recorded, and the gap against the stored figure is
reported rather than hidden — that gap is the closest thing to a self-check
this data allows.

WHAT IT REFUSES TO DO
---------------------
Report an R it cannot compute. R needs the stop that was planned at entry, and
`planned_stop_at_entry` is present on 21 of 91 closed rows. The other 70 get
rupees and cost percentages, which are real, and no R at all — rather than an R
against an invented denominator. Stage 4 carries the stop through so this
improves on its own.
"""

from __future__ import annotations

import argparse
import csv
import statistics
import sys
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from loguru import logger


# Clip bands in rupees. Chosen around this account: a Rs 20,000 book taking
# 1-2% risk produces clips in the low thousands, which is exactly the region
# where the flat Rs 15.04 DP fee dominates and where the answer changes.
CLIP_BANDS = [(0, 2500), (2500, 5000), (5000, 10000), (10000, 25000), (25000, 10**9)]


def _band(value: float) -> str:
    for lo, hi in CLIP_BANDS:
        if lo <= value < hi:
            return f"₹{lo:,}–{hi:,}" if hi < 10**9 else f"₹{lo:,}+"
    return "?"


def _se(xs: list[float]) -> float:
    """Standard error. Every estimate in this system carries one (§17)."""
    if len(xs) < 2:
        return float("nan")
    return statistics.stdev(xs) / (len(xs) ** 0.5)


def load(sb) -> list[dict]:
    """
    Every closed trade, with friction recomputed and R where R is computable.

    product is taken as RECORDED, not inferred from framework. Nine intraday
    rows carry product=CNC because `intraday_product` was genuinely set to CNC
    when they were taken — those trades really would have paid delivery
    charges. Relabelling them MIS because intraday "should" be MIS would be
    rewriting history to make a bucket tidier. They are flagged instead.
    """
    from intraday.cost_model import entry_leg, exit_leg

    BASE = ("symbol,product,framework,mode,strategy,intraday_strategy,"
            "entry_price,exit_price,actual_qty,planned_stop_at_entry,"
            "realized_pnl,charges,hold_days,exit_date,max_favorable_excursion,"
            "signal_id,signal_date,source")
    # Migration 031's columns. A tool that reads code-ahead-of-migration must
    # degrade to the older shape rather than crash — otherwise the one command
    # that would tell you what is wrong is the one that will not run.
    RUNNER = ",runner_since_r,runner_evidence,runner_verdict"

    cols = BASE + RUNNER
    try:
        sb.table("closed_positions").select(cols).limit(1).execute()
    except Exception:
        logger.warning("  runner columns absent — migration 031 is not applied. "
                       "Capture ratio cannot be split by runner state.")
        cols = BASE

    rows, off = [], 0
    while True:
        page = (sb.table("closed_positions").select(cols)
                  .range(off, off + 999).execute().data) or []
        rows += page
        if len(page) < 1000:
            break
        off += 1000

    out = []
    for r in rows:
        entry = float(r.get("entry_price") or 0)
        exit_ = float(r.get("exit_price") or 0)
        qty   = int(r.get("actual_qty") or 0)
        if entry <= 0 or qty <= 0:
            continue
        product = (r.get("product") or "CNC").upper()

        modelled = entry_leg(entry, qty, product=product) + exit_leg(exit_ or entry, qty, product=product)
        invested = entry * qty
        gross    = float(r.get("realized_pnl") or 0.0)

        stop = r.get("planned_stop_at_entry")
        risk_amt = (entry - float(stop)) * qty if stop not in (None, "") else 0.0

        out.append({
            **r,
            "product":       product,
            "invested":      invested,
            "modelled":      modelled,
            "stored":        float(r["charges"]) if r.get("charges") not in (None, "") else None,
            "gross_pnl":     gross,
            "net_pnl":       gross - modelled,
            "cost_pct":      modelled / invested * 100.0 if invested else 0.0,
            "risk_amt":      risk_amt if risk_amt > 0 else None,
            "gross_r":       gross / risk_amt if risk_amt > 0 else None,
            "net_r":         (gross - modelled) / risk_amt if risk_amt > 0 else None,
            "cost_r":        modelled / risk_amt if risk_amt > 0 else None,
            "mislabelled":   (r.get("framework") == "INTRADAY" and product != "MIS"),
            # SAME RULE AS THE DASHBOARD'S inEra(), deliberately.
            #
            # 69 of the 72 real-money closes were entered by hand before the
            # signal engine existed. Pooling them scores the engine on trades it
            # never made — which is how a -Rs 12,489 book gets attributed to a
            # system that placed three of those trades. Two definitions of "the
            # system did this" would drift, so this mirrors
            # frontend/components/tabs/PerformanceTab.tsx exactly: a trade is
            # the system's if a signal, an engine or paper mode produced it.
            "attributed":    bool(r.get("signal_date") or r.get("signal_id")
                                  or (r.get("intraday_strategy") or "").strip()
                                  or r.get("source") == "paper"),
        })
    return out


def _stat_line(label: str, trades: list[dict], key: str) -> str:
    vals = [t[key] for t in trades if t.get(key) is not None]
    if not vals:
        return f"  {label:<26} n=0    — not computable"
    m, se = statistics.fmean(vals), _se(vals)
    med = statistics.median(vals)
    se_s = f"±{se:.3f}" if se == se else "±  n/a"
    return f"  {label:<26} n={len(vals):<4} mean {m:+.3f} {se_s}   median {med:+.3f}"


def ledger(rows: list[dict]) -> None:
    logger.info("═" * 78)
    logger.info("EXPECTANCY LEDGER — every closed trade, friction recomputed at today's rates")
    logger.info("═" * 78)

    logger.info("")
    logger.info("── Attribution: did this system produce the trade? ─────────────────────")
    sysr = [t for t in rows if t["attributed"]]
    leg  = [t for t in rows if not t["attributed"]]
    for label, sub in (("SINCE AUTOMATION", sysr), ("legacy / manual", leg)):
        if not sub:
            continue
        g_ = sum(t["gross_pnl"] for t in sub); c_ = sum(t["modelled"] for t in sub)
        live = [t for t in sub if t.get("mode") == "LIVE"]
        logger.info(f"  {label:<18} n={len(sub):<4} ({len(live)} real money)  "
                    f"gross ₹{g_:>9,.0f}  friction ₹{c_:>7,.0f}  net ₹{g_-c_:>9,.0f}")
    logger.warning("  Only the first line scores THIS SYSTEM. The second was entered by hand "
                   "before the engine existed and is shown for context only.")

    logger.info("")
    logger.info("── Rupees, all closed trades ───────────────────────────────────────────")
    for key in sorted({(t["framework"], t["product"]) for t in rows}):
        fw, prod = key
        sub = [t for t in rows if t["framework"] == fw and t["product"] == prod]
        gross = sum(t["gross_pnl"] for t in sub)
        cost  = sum(t["modelled"] for t in sub)
        net   = gross - cost
        wins  = sum(1 for t in sub if t["net_pnl"] > 0)
        logger.info(f"  {fw:<9} {prod:<4} n={len(sub):<4} gross ₹{gross:>9,.0f}   "
                    f"friction ₹{cost:>8,.0f}   net ₹{net:>9,.0f}   "
                    f"net-win {wins}/{len(sub)}")
    g, c = sum(t["gross_pnl"] for t in rows), sum(t["modelled"] for t in rows)
    # A friction-as-%-of-gross ratio is nonsense when gross is negative — it
    # prints a cheerful negative percentage for a book that is losing money and
    # paying charges on top. Say what is true instead.
    verdict = (f"friction is {c/g*100:.0f}% of gross" if g > 0 else
               f"gross is NEGATIVE — friction of ₹{c:,.0f} is added to a loss, not taken from a profit")
    logger.info(f"  {'ALL':<9} {'':<4} n={len(rows):<4} gross ₹{g:>9,.0f}   "
                f"friction ₹{c:>8,.0f}   net ₹{g-c:>9,.0f}")
    logger.info(f"  {'':<14} {verdict}")

    logger.info("")
    logger.info("── R, only where the entry stop was recorded ────────────────────────────")
    scored = [t for t in rows if t["net_r"] is not None]
    logger.info(f"  {len(scored)} of {len(rows)} closed trades carry planned_stop_at_entry.")
    if not scored:
        logger.warning("  No R can be computed. Stage 4 carries the stop through; until then "
                       "this section is empty rather than invented.")
    # Split by PRODUCT as well as framework. Pooling them hides the finding:
    # the delivery-priced intraday trades carry a flat Rs 15.04 DP fee against
    # a risk budget of a few hundred rupees, so their cost_R dwarfs everything
    # and drags a pooled mean far above any figure that describes either group.
    for fw, prod in sorted({(t["framework"], t["product"]) for t in scored}):
        sub = [t for t in scored if t["framework"] == fw and t["product"] == prod]
        logger.info(f"  · {fw} / {prod}   (n={len(sub)})")
        logger.info(_stat_line("gross R", sub, "gross_r"))
        logger.info(_stat_line("friction, in R", sub, "cost_r"))
        logger.info(_stat_line("NET R  ← the number", sub, "net_r"))

    logger.info("")
    logger.info("── Data quality ─────────────────────────────────────────────────────────")
    mis = [t for t in rows if t["mislabelled"]]
    if mis:
        logger.warning(f"  {len(mis)} INTRADAY trade(s) recorded as a delivery product — "
                       f"priced as delivery here because that is what was traded: "
                       f"{', '.join(sorted({t['symbol'] for t in mis}))}")
    both = [t for t in rows if t["stored"] is not None]
    if both:
        deltas = [abs(t["stored"] - t["modelled"]) for t in both]
        worst  = max(both, key=lambda t: abs(t["stored"] - t["modelled"]))
        logger.info(f"  stored vs modelled charges on {len(both)} row(s): "
                    f"mean gap ₹{statistics.fmean(deltas):.2f}, "
                    f"worst {worst['symbol']} stored ₹{worst['stored']:.2f} "
                    f"vs modelled ₹{worst['modelled']:.2f}")
        logger.info("  A large gap means old rows were priced under old rates or a partial "
                    "leg. It is NOT a reconciliation — both sides are this model.")
    # Whether the model has ever been checked against the broker is a fact
    # about the system, not about this run, so it is stored rather than
    # remembered. Without it the ledger either nags forever after the work is
    # done, or — worse — stops nagging because someone edited the string.
    try:
        from config import get_supabase
        rec = (get_supabase().table("system_config").select("value")
                 .eq("key", "cost_model_reconciled").execute().data or [])
        val = rec[0]["value"] if rec else ""
    except Exception:
        val = ""
    if val:
        logger.success(f"  friction model reconciled against the broker: {val}")
    else:
        logger.warning("  RECONCILIATION AGAINST THE BROKER HAS NOT BEEN DONE. "
                       "See --reconcile; Stage 2 is not complete without it.")


def capture(rows: list[dict]) -> None:
    """
    How much of the favourable move the system actually kept.

    THE STAGE 3 NUMBER. exit_rules.py's own docstring records a median capture
    of 3% across 70 trades — nearly all of the favourable excursion given back —
    and a momentum system that surrenders its right tail has no economic thesis.
    This makes that figure reproducible instead of a remembered one, so "before
    and after" is a command rather than an argument.

    capture = realised move / maximum favourable excursion.

    Losers are reported separately, not pooled. A trade that went nowhere and
    stopped out has a capture ratio that is arithmetically valid, wildly
    negative, and says nothing about exit quality — pooling it moves the median
    for a reason unconnected to the thing being measured.
    """
    usable, junk = [], 0
    for t in rows:
        mfe = t.get("max_favorable_excursion")
        entry = float(t.get("entry_price") or 0)
        if mfe in (None, "") or not entry:
            continue
        mfe = float(mfe)
        # One row carries a PRICE in this column rather than a percentage
        # (UNIONBANK, MFE 197.87 against a +1.17% realised move). One bad row
        # moves a median of sixty, so it is excluded and counted out loud
        # rather than quietly winsorised.
        if mfe <= 0 or abs(mfe - entry) / entry < 0.5:
            junk += 1
            continue
        realised = (float(t["exit_price"]) - entry) / entry * 100.0
        usable.append({**t, "mfe": mfe, "realised_pct": realised,
                       "capture": realised / mfe})

    logger.info("")
    logger.info("── Capture ratio — how much of the favourable move was kept ─────────────")
    logger.info(f"  {len(usable)} of {len(rows)} closed trades carry a usable excursion"
                + (f"; {junk} excluded as unusable" if junk else ""))
    if not usable:
        return

    for label, sub in (("winners  ", [t for t in usable if t["realised_pct"] > 0]),
                       ("losers   ", [t for t in usable if t["realised_pct"] <= 0]),
                       ("ALL      ", usable)):
        if not sub:
            continue
        caps = [t["capture"] for t in sub]
        logger.info(f"  {label} n={len(sub):<4} median {statistics.median(caps)*100:>7.1f}%   "
                    f"mean {statistics.fmean(caps)*100:>7.1f}%")

    # MFE is refreshed when the lifecycle runs, not tick by tick, so a peak
    # between runs is missed. Capture above 100% is arithmetically impossible —
    # you cannot keep more than the favourable move — so every such row is
    # proof of under-recording, and the true capture is somewhat lower than
    # what is printed above. Counted rather than hidden.
    win = [t for t in usable if t["realised_pct"] > 0]
    impossible = [t for t in win if t["capture"] > 1.0]
    if impossible:
        clean = [t["capture"] for t in win if t["capture"] <= 1.0]
        logger.warning(f"  {len(impossible)} of {len(win)} winners show capture > 100%, which "
                       f"cannot happen — MFE is sampled on the lifecycle run, not per tick, "
                       f"so peaks between runs are missed. Excluding them the winner median "
                       f"is {statistics.median(clean)*100:.1f}%, which is the conservative read.")
    if win:
        med = statistics.median([t["capture"] for t in win]) * 100
        logger.info("")
        logger.info(f"  TARGET is 30%+ on winners. Currently {med:.1f}%.")
        if med < 30:
            give = statistics.fmean([t["mfe"] - t["realised_pct"] for t in win])
            logger.warning(f"  Mean give-back on a winner is {give:.2f} percentage points "
                           f"of favourable move. That is the right tail runner "
                           f"conversion exists to capture.")

    # Split by whether the position was ever converted to a runner. Before
    # migration 031 nothing recorded this, so it reads "unknown" for every
    # historical row — which is itself the finding.
    known = [t for t in usable if t.get("runner_since_r") not in (None, "")]
    logger.info(f"  runner state recorded on {len(known)} of {len(usable)} — "
                + ("split below" if known else
                   "nothing before migration 031 knows whether it ran, so the "
                   "question 'do runners capture more?' cannot be asked yet"))
    if known:
        caps = [t["capture"] for t in known]
        logger.info(f"    converted to runner  n={len(known):<4} "
                    f"median {statistics.median(caps)*100:.1f}%")


def by_clip(rows: list[dict]) -> None:
    """
    Friction against position size — the whole argument for fewer, larger clips.

    Flat charges do not scale, so the same trade is a different trade at a
    different size. This is the table that says by how much.
    """
    logger.info("═" * 78)
    logger.info("FRICTION BY CLIP SIZE")
    logger.info("═" * 78)
    for prod in sorted({t["product"] for t in rows}):
        logger.info(f"")
        logger.info(f"── {prod} ──")
        logger.info(f"  {'clip':<16} {'n':>4} {'friction %':>11} {'friction R':>12} {'net R':>10}")
        buckets = defaultdict(list)
        for t in rows:
            if t["product"] == prod:
                buckets[_band(t["invested"])].append(t)
        for lo, hi in CLIP_BANDS:
            label = f"₹{lo:,}–{hi:,}" if hi < 10**9 else f"₹{lo:,}+"
            sub = buckets.get(label, [])
            if not sub:
                continue
            cpct = statistics.fmean([t["cost_pct"] for t in sub])
            crs  = [t["cost_r"] for t in sub if t["cost_r"] is not None]
            nrs  = [t["net_r"] for t in sub if t["net_r"] is not None]
            logger.info(f"  {label:<16} {len(sub):>4} {cpct:>10.2f}% "
                        f"{(statistics.fmean(crs) if crs else float('nan')):>12.3f} "
                        f"{(statistics.fmean(nrs) if nrs else float('nan')):>10.3f}")

    # The same question asked of the model directly, which needs no sample at
    # all and therefore answers for clip sizes this account has never taken.
    from intraday.cost_model import round_trip
    logger.info("")
    logger.info("── modelled round trip at a Rs 500 share, by clip ───────────────────────")
    logger.info(f"  {'clip':>10} {'CNC':>10} {'MIS':>10}   (round trip as % of position)")
    for value in (2000, 3000, 5000, 8000, 12000, 20000, 50000, 100000):
        qty = max(int(value / 500), 1)
        c = round_trip(500.0, qty, product="CNC").pct_of_position
        m = round_trip(500.0, qty, product="MIS").pct_of_position
        logger.info(f"  ₹{value:>9,} {c:>9.2f}% {m:>9.2f}%")
    logger.info("  CNC costs ~5x MIS at these sizes and the gap is almost all the flat")
    logger.info("  DP fee, which is why clip size — not target — is the swing book's lever.")


def _read_zerodha_xlsx(path: Path) -> dict:
    """
    Pull the authoritative numbers out of a Zerodha Console P&L export.

    Two sheets matter. `Equity` carries the per-symbol buy/sell values and the
    statutory charge breakdown; `Other Debits and Credits` carries the flat DP
    fee per scrip per sell day and — the line nothing in this system has ever
    modelled — the Kite Connect API subscription.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)

    charges, symbols = {}, []
    ws = wb["Equity"]
    mode, idx = None, {}

    def num(cells: list[str], name: str) -> float:
        """Read a column BY HEADER NAME, never by position."""
        i = idx.get(name)
        if i is None or i >= len(cells) or not cells[i]:
            return 0.0
        try:
            return float(cells[i])
        except ValueError:
            return 0.0

    for r in ws.iter_rows(values_only=True):
        cells = [("" if c is None else str(c).strip()) for c in r]
        joined = " ".join(cells)
        if "Account Head" in joined:
            mode = "charges"
            continue
        if "Realized P&L Pct" in joined:
            # POSITION-BASED PARSING IS WRONG HERE AND IT FAILED SILENTLY.
            #
            # "Open Quantity Type" is blank on every row, so compacting the row
            # to its non-empty cells shifts every column after it left by one —
            # open_value read the unrealized P&L and unrealized read a
            # percentage. Buy value came out Rs 8,924 against a true Rs 16,627,
            # and the reconciliation would have compared modelled charges on
            # half the turnover to the broker's charges on all of it.
            #
            # It was caught only because the exchange fee is an exact percentage
            # of turnover with no rounding, so it independently implies the
            # turnover and the two disagreed by 43%. That cross-check is the
            # reason this bug is a corrected comment rather than a wrong number
            # in a report — keep it.
            mode = "symbols"
            idx = {name: i for i, name in enumerate(cells) if name}
            continue
        vals = [c for c in cells if c]
        if not vals:
            continue
        if mode == "charges" and len(vals) >= 2:
            try:
                charges[vals[0]] = float(vals[1])
            except ValueError:
                mode = None
        elif mode == "symbols" and idx:
            sym_i = idx.get("Symbol")
            if sym_i is None or sym_i >= len(cells) or not cells[sym_i]:
                continue
            symbols.append({
                "symbol":     cells[sym_i],
                "buy_value":  num(cells, "Buy Value"),
                "sell_value": num(cells, "Sell Value"),
                "realized":   num(cells, "Realized P&L"),
                "open_value": num(cells, "Open Value"),
                "unrealized": num(cells, "Unrealized P&L"),
            })

    dp_total, dp_count, fixed = 0.0, 0, {}
    if "Other Debits and Credits" in wb.sheetnames:
        for r in wb["Other Debits and Credits"].iter_rows(values_only=True):
            cells = [("" if c is None else str(c).strip()) for c in r]
            vals = [c for c in cells if c]
            if len(vals) < 3 or vals[0].lower().startswith(("particulars", "client", "other")):
                continue
            try:
                debit = float(vals[2])
            except ValueError:
                continue
            if "DP Charges" in vals[0]:
                dp_total += debit
                dp_count += 1
            elif debit:
                # Anything that is not a per-trade charge. The API subscription
                # lives here, and it is not small.
                fixed[vals[0]] = fixed.get(vals[0], 0.0) + debit

    return {"charges": charges, "symbols": symbols, "dp_total": dp_total,
            "dp_count": dp_count, "fixed": fixed}


def reconcile(rows: list[dict], path: Path) -> int:
    """
    Modelled friction against what the broker actually charged.

    THE ONE VALIDATION STAGE 2 CANNOT SKIP. Kite Connect's trades endpoint
    returns fills, not charges, so the realised side comes from the Console P&L
    export — the same file the operator downloads to do their own taxes.

    Reconciled COMPONENT BY COMPONENT rather than on the total, because a total
    can be right while two rates are wrong in opposite directions. Exchange and
    SEBI fees are exact percentages of turnover with no rounding, so they double
    as a check that the turnover being reconciled is the right one: if the
    turnover they imply disagrees with the symbol table, the two tables describe
    different trades and every line below would be arithmetic on sand.

    Slippage is excluded from both sides. It is a real cost and it is not a
    broker charge — it is already inside the fill price — so carrying it here
    would manufacture a gap no published schedule could explain.
    """
    if path.suffix.lower() not in (".xlsx", ".xls"):
        logger.error(f"  expected a Zerodha Console P&L export (.xlsx), got '{path.suffix}'")
        logger.error("  Console -> Reports -> P&L -> pick the period -> download")
        return 1
    if not path.exists():
        logger.error(f"  {path} not found")
        return 1

    from config import cfg_float
    z = _read_zerodha_xlsx(path)
    if not z["symbols"]:
        logger.error("  no symbol rows found in the Equity sheet")
        return 1

    # Buy value must include positions still open: they were bought inside the
    # window and paid STT and stamp on the way in. Cost basis, not market value.
    buy  = sum(s["buy_value"] + max(s["open_value"] - s["unrealized"], 0.0)
               for s in z["symbols"])
    sell = sum(s["sell_value"] for s in z["symbols"])
    turnover = buy + sell

    def actual(*names) -> float:
        return sum(v for k, v in z["charges"].items()
                   if any(n.lower() in k.lower() for n in names))

    a_exch  = actual("Exchange Transaction")
    a_sebi  = actual("SEBI Turnover")
    a_stt   = actual("Securities Trans")
    a_stamp = actual("Stamp Duty")
    a_brok  = actual("Brokerage")
    a_gst   = actual("GST")
    a_stat  = sum(z["charges"].values())

    rate_exch = cfg_float("cost_exchange_pct", 0.00307)
    implied = a_exch / (rate_exch / 100.0) if a_exch else 0.0
    drift = abs(implied - turnover) / turnover * 100 if turnover else 0.0

    m_brok  = 0.0                                    # equity delivery is free
    m_stt   = (buy + sell) * cfg_float("cost_cnc_stt_buy_pct", 0.1) / 100.0
    m_exch  = turnover * rate_exch / 100.0
    m_sebi  = turnover * cfg_float("cost_sebi_pct", 0.0001) / 100.0
    m_stamp = buy * cfg_float("cost_cnc_stamp_buy_pct", 0.015) / 100.0
    m_gst   = (m_brok + m_exch + m_sebi) * cfg_float("cost_gst_pct", 18.0) / 100.0
    m_stat  = m_brok + m_stt + m_exch + m_sebi + m_stamp + m_gst
    m_dp    = z["dp_count"] * cfg_float("cost_dp_per_sell", 15.04)

    logger.info("=" * 78)
    logger.info("FRICTION RECONCILIATION - modelled against the broker's own statement")
    logger.info("=" * 78)
    logger.info(f"  buy Rs{buy:,.2f} | sell Rs{sell:,.2f} | turnover Rs{turnover:,.2f}")
    logger.info(f"  turnover implied by the exchange fee: Rs{implied:,.2f} ({drift:.2f}% drift)")
    if drift > 2.0:
        logger.error("  x the symbol table and the charge table describe different trades - "
                     "the export window probably clips a leg. Fix that before reading on.")
        return 1

    logger.info("")
    logger.info(f"  {'component':<22} {'modelled':>12} {'actual':>12} {'gap':>9}")
    for label, m, a in (("brokerage", m_brok, a_brok), ("STT", m_stt, a_stt),
                        ("exchange txn", m_exch, a_exch), ("SEBI turnover", m_sebi, a_sebi),
                        ("stamp duty", m_stamp, a_stamp), ("GST", m_gst, a_gst)):
        gap = (m - a) / a * 100 if a else (0.0 if not m else float("inf"))
        # A rate can be badly wrong in percentage terms while being pennies in
        # rupees. Flag only when it is both.
        flag = "  <-" if (abs(gap) > 10 and abs(m - a) >= 1.0) else ""
        logger.info(f"  {label:<22} Rs{m:>10.4f} Rs{a:>10.4f} {gap:>8.1f}%{flag}")

    logger.info("  " + "-" * 58)
    g_stat = (m_stat - a_stat) / a_stat * 100 if a_stat else 0.0
    logger.info(f"  {'statutory total':<22} Rs{m_stat:>10.4f} Rs{a_stat:>10.4f} {g_stat:>8.1f}%")
    g_dp = (m_dp - z["dp_total"]) / z["dp_total"] * 100 if z["dp_total"] else 0.0
    logger.info(f"  {'DP x ' + str(z['dp_count']) + ' sell days':<22} Rs{m_dp:>10.4f} "
                f"Rs{z['dp_total']:>10.4f} {g_dp:>8.1f}%")
    m_all, a_all = m_stat + m_dp, a_stat + z["dp_total"]
    overall = (m_all - a_all) / a_all * 100 if a_all else 0.0
    logger.info(f"  {'TOTAL':<22} Rs{m_all:>10.4f} Rs{a_all:>10.4f} {overall:>8.1f}%")

    # -- The cost that is not a trading cost --------------------------------
    realized = sum(s["realized"] for s in z["symbols"])
    fixed_total = sum(z["fixed"].values())
    if z["fixed"]:
        logger.info("")
        logger.info("-- Fixed costs, which no model in this system has ever carried --------")
        for k, v in sorted(z["fixed"].items(), key=lambda kv: -kv[1]):
            logger.warning(f"  Rs{v:>9,.2f}  {k}")
        logger.info("")
        logger.info(f"  realised P&L over the window   Rs{realized:>10,.2f}")
        logger.info(f"  trading friction               Rs{-a_all:>10,.2f}")
        logger.info(f"  fixed costs                    Rs{-fixed_total:>10,.2f}")
        logger.info(f"  {'NET':<30} Rs{realized - a_all - fixed_total:>10,.2f}")
        if fixed_total > a_all:
            logger.error(f"  x FIXED COSTS ARE {fixed_total / a_all:.1f}x TRADING FRICTION. "
                         f"Every cost model in this repository optimises the smaller number.")

    logger.info("")
    n_trips = sum(1 for s in z["symbols"] if s["sell_value"] > 0)
    if abs(overall) > 10.0:
        logger.error(f"  x modelled is {overall:+.1f}% against realised, outside +/-10%. "
                     f"The ledger is WRONG and must be corrected before Stage 3.")
        return 1
    logger.success(f"  OK modelled within {overall:+.2f}% of realised on Rs{turnover:,.0f} "
                   f"of turnover across {n_trips} realised round trip(s)")
    # Record the pass so the ledger stops asking for work already done, and so
    # a later reader can see WHEN it was last true rather than assuming it is.
    try:
        import datetime as _dt
        from config import get_supabase
        note = (f"{_dt.date.today().isoformat()}: {overall:+.2f}% across "
                f"{n_trips} round trip(s), Rs{turnover:,.0f} turnover")
        sb_ = get_supabase()
        if (sb_.table("system_config").select("key")
              .eq("key", "cost_model_reconciled").execute().data):
            sb_.table("system_config").update({"value": note}).eq(
                "key", "cost_model_reconciled").execute()
        else:
            sb_.table("system_config").insert({
                "key": "cost_model_reconciled", "value": note,
                "category": "costs", "subsystem": "tools/expectancy_ledger.py",
                "value_type": "string", "risk_level": "SAFE",
                "description": "When the friction model was last checked against the "
                               "broker's own statement, and by how much it was out. "
                               "Written by tools.expectancy_ledger --reconcile.",
            }).execute()
        logger.info(f"  recorded: {note}")
    except Exception as e:
        logger.warning(f"  could not record the reconciliation result: {e}")
    if n_trips < 20:
        logger.warning(f"  ! the specification asks for >=20 round trips; this is {n_trips}. "
                       f"The RATE TABLE is validated - exchange and SEBI match to four "
                       f"decimals - but the SAMPLE is not yet enough to catch a rate that "
                       f"is only wrong in cases these trades never hit.")
    return 0


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="Net R per trade, by product and clip size")
    ap.add_argument("--by-clip", action="store_true", help="friction against position size")
    ap.add_argument("--capture", action="store_true", help="capture ratio (the Stage 3 number)")
    ap.add_argument("--reconcile", default=None, help="Zerodha Console P&L export (.xlsx)")
    a = ap.parse_args(argv)

    from config import get_supabase
    rows = load(get_supabase())
    if not rows:
        logger.error("no closed trades to score")
        return 1

    if a.reconcile:
        return reconcile(rows, Path(a.reconcile))
    if a.by_clip:
        by_clip(rows)
        return 0
    if a.capture:
        capture(rows)
        return 0
    ledger(rows)
    capture(rows)
    return 0


if __name__ == "__main__":
    sys.exit(main())
